import pyodbc
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
import tkinter as tk
import os

def generate_nokia_report(iso_year, iso_week, vendor_no):
    # Create the generating report popup
    generating_report_popup = tk.Toplevel(window)
    generating_report_popup.geometry("250x50")
    generating_report_popup.title("Generating Report")
    generating_report_popup.transient(window)
    generating_report_popup.grab_set()

    generating_report_label = tk.Label(generating_report_popup, text="Generating report...")
    generating_report_label.pack(pady=10)


    # calculate the start and end dates of the specified week
    iso_date = datetime.strptime(f'{iso_year}-W{iso_week}-1', '%G-W%V-%u').date()
    start_date = iso_date - timedelta(days=iso_date.weekday())
    iso_end_date = start_date + timedelta(days=6)
    transaction_end_date = datetime.today().date()

    # connect to the database using Windows authentication
    server = 'jeg-psql2'
    database = 'JEGSONS-NAV60-PRO'
    cnxn = pyodbc.connect(
        f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes')
    cursor = cnxn.cursor()

    # SQL QUERIES-----------------------------------------------------------------------------------------------------------

    # Query Item Ledger Entry Table - use a parameterized query to substitute variables in the SQL query
    ILE_query = "SELECT [Item No_], [Posting Date], [Entry Type], [Source No_], [Document No_], [Location Code], [Quantity], [No_ Series], [Document Type]   FROM [JEG_SONS, Inc_$Item Ledger Entry] WHERE [Posting Date] >= ? AND [Posting Date] <= ?"
    transactions = pd.read_sql(ILE_query, cnxn, params=[start_date, transaction_end_date])
    transactions['Posting Date'] = transactions['Posting Date'].dt.date


    # Query Item Table - use a parameterized query to substitute variables in the SQL query
    Item_query = "SELECT [No_], [Description], [Vendor No_], [Vendor Item No_], [Inventory Available], [Product Group Code], [Item Category Code] FROM [JEG_SONS, Inc_$Item] WHERE [Vendor No_] = ?"
    items = pd.read_sql(Item_query, cnxn, params=[vendor_no])

    # Filter on Nokia new items only
    items_nokia = items[(items['Vendor No_'] == vendor_no) & (items['Vendor Item No_'] != '') & (items['Product Group Code'] == 'NOKIA')
                                & (items['Item Category Code'] == 'PHONE-NEW')]
    # items_nokia.to_excel(r'items nokia.xlsx', index=False)


    # Query PO Lines Table - use a parameterized query to substitute variables in the SQL query
    PO_query = "SELECT [Document Type], [Buy-from Vendor No_], [Type], [No_], [Outstanding Quantity], [Order Date] FROM [JEG_SONS, Inc_$Purchase Line] WHERE [Document Type] = 1 AND [Buy-from Vendor No_] = ? AND  [Type] = 2 AND [Order Date] <= ?"
    PO_lines = pd.read_sql(PO_query, cnxn, params=[vendor_no,iso_end_date])

    # Query Customer Table
    Customer_query = "SELECT [No_], [Name], [City], [Country_Region Code] FROM [JEG_SONS, Inc_$Customer]"
    Customers = pd.read_sql(Customer_query, cnxn)

    # Query Posted Package Line Table
    item_nokia_list = items_nokia['No_'].tolist()

    Package_query = "SELECT [No_], [Serial No_], [Packing Date], [Type], [Source Type], [Source Subtype], [Posted Source ID] FROM [JEG_SONS, Inc_$Posted Package Line] WHERE [Type] = 2 AND [Source Type] = 36 AND [Source Subtype] = 1 AND [Packing Date] >= ? AND [Packing Date] <= ? AND [No_] IN ({})".format(','.join('?'*len(item_nokia_list)))
    params = [start_date, iso_end_date] + item_nokia_list
    Package_lines = pd.read_sql(Package_query, cnxn, params=params)


    # Prepare Data for Summary Tab -----------------------------------------------------------------------------------------

    # merge the items dataframe with the transactions dataframe
    transactions = pd.merge(transactions, items, how='left', left_on='Item No_', right_on='No_')
    transactions = transactions.drop('No_', axis=1)

    # create the blank running total column
    transactions['Running Total'] = 0

    # loop through the transactions and update running total column for each row
    for i, row in transactions.iterrows():
        current_item = row['Item No_']
        current_qty = row['Quantity']
        current_inventory = row['Inventory Available']

        # grab all transactions after current transaction and sum up the quantity to reverse
        after_transactions = transactions[i + 1:]
        after_transactions = after_transactions[after_transactions['Item No_'] == current_item]
        after_transactions['Quantity'] = -1 * after_transactions['Quantity']
        qty_to_reverse = after_transactions['Quantity'].sum()

        # update running total
        running_total = current_inventory + qty_to_reverse
        transactions.at[i, 'Running Total'] = running_total

    # For Testing Purposes
    # transactions.to_excel(r'transactions.xlsx', index=False)

    transactions_in_week = transactions[(transactions['Vendor No_'] == vendor_no) & (transactions['Posting Date'] <= iso_end_date)
                                & (transactions['Vendor Item No_'].notnull()) & (transactions['Product Group Code'] == 'NOKIA')
                                & (transactions['Item Category Code'] == 'PHONE-NEW')]
    # transactions_in_week.to_excel(r'transactions_in_week.xlsx', index=False)


    # create a list to hold the data for each item
    summary_data = []

    # loop through the items and run calculations for each item
    for item, row in items_nokia.iterrows():
        inventory_available = row['Inventory Available']

        opening_inventory = 0
        if len(transactions.loc[transactions['Item No_'] == row['No_']]) > 0:
            opening_transaction = transactions.loc[transactions['Item No_'] == row['No_'], ['Quantity', 'Running Total']].iloc[0]
            opening_inventory = opening_transaction['Running Total'] - opening_transaction['Quantity']
        else:
            opening_inventory = row['Inventory Available']

        goods_received = transactions_in_week.loc[(transactions_in_week['Item No_'] == row['No_']) & (transactions_in_week['Entry Type'] == 0) & (transactions_in_week['Document Type'] == 5), 'Quantity'].sum()
        sell_thru = -1 * transactions_in_week.loc[(transactions_in_week['Item No_'] == row['No_']) & (transactions_in_week['Entry Type'] == 1) & (transactions_in_week['Document Type'] == 1), 'Quantity'].sum()
        adjustment = transactions_in_week.loc[(transactions_in_week['Item No_'] == row['No_']) & (
                    (transactions_in_week['Entry Type'] == 2) | (
                        transactions_in_week['Entry Type'] == 3)), 'Quantity'].sum()

        credit_memos = transactions_in_week.loc[(transactions_in_week['Item No_'] == row['No_']) & (
                transactions_in_week['Entry Type'] == 1) & (transactions_in_week['Document Type'] == 3), 'Quantity'].sum()

        adjustment = adjustment + credit_memos

        current_in_transit = PO_lines[PO_lines['No_'] == row['No_']]['Outstanding Quantity'].sum() if len(PO_lines) > 0 else 0

        # get the warehouse for the item
        warehouse = 0
        if len(transactions_in_week.loc[transactions_in_week['Item No_'] == row['No_']]) > 0:
            warehouse = transactions_in_week.loc[transactions_in_week['Item No_'] == row['No_'], 'Running Total'].iloc[-1]
        else:
            warehouse = 0

        closing_inventory = warehouse + current_in_transit

        # create a dictionary with the data for this item
        item_dict = {
            'Week': f"Week {iso_week}",
            'Sales Pack': row['Description'],
            'Product Code': row['Vendor Item No_'],
            'Opening inventory': opening_inventory,
            'Goods Received': goods_received,
            'Sell thru': sell_thru,
            'Adjustment': adjustment,
            'Closing inventory': closing_inventory,
            'In-transit': current_in_transit,
            'Warehouse': warehouse
        }

        # add the dictionary to the list of item data
        summary_data.append(item_dict)

    # create a dataframe from the list of dictionaries
    df = pd.DataFrame(summary_data)

    # Prepare data for Sales breakdown tab----------------------------------------------------------------------------------

    # Group Shipments By Vendor Item No and Source No, then sum Quantities
    shipments_in_week = transactions_in_week[(transactions_in_week['No_ Series'] == 'S-SHPT')]

    grouped_transactions = shipments_in_week.groupby(['Vendor Item No_', 'Source No_'])['Quantity'].sum().reset_index()
    grouped_transactions = pd.merge(grouped_transactions, items_nokia[['Vendor Item No_', 'Description']], on='Vendor Item No_', how='left')


    # Merge grouped transactions with Customers query
    transactions_with_cust_info = pd.merge(grouped_transactions, Customers, how='left', left_on='Source No_', right_on='No_')
    transactions_with_cust_info.drop('No_', axis=1, inplace=True)  # drop redundant column
    # transactions_with_cust_info.to_excel(r'sales breakdown.xlsx', index=False)

    sales_breakdown_data = []

    for shipments, row in transactions_with_cust_info.iterrows():
        sales_dict = {
                'Week': f"Week {iso_week}",
                'POS-ID': row['Source No_'],
                'POS Name': row['Name'],
                'City': row['City'],
                'Country': row['Country_Region Code'],
                'Model': row['Description'],
                'Product Code': row['Vendor Item No_'],
                'Volume Sales': -1 * row['Quantity']
                }

        sales_breakdown_data.append(sales_dict)

    df2 = pd.DataFrame(sales_breakdown_data)


    # Prepare data for IMEI Sell thru tab-----------------------------------------------------------------------------------

    Package_lines = pd.merge(Package_lines,shipments_in_week[['Document No_', 'Source No_']], how='left', left_on='Posted Source ID', right_on='Document No_' )
    Package_lines = Package_lines.drop(['Document No_'], axis=1)

    Customers_renamed = Customers.rename(columns={'No_': 'No_Customer'})

    Package_lines = pd.merge(Package_lines,Customers_renamed[['No_Customer', 'Name']], how='left', left_on='Source No_', right_on='No_Customer')
    Package_lines = Package_lines.drop(['No_Customer'], axis=1)
    Package_lines['Name'].replace("RETURN CENTER", "WALMART DROPSHIP", inplace=True)
    # Package_lines.to_excel("Packages_in_week.xlsx", index=False)

    imei_data = []

    for shipments, row in Package_lines.iterrows():
        imei_dict = {
                'Retail store name': row['Name'],
                'Retail store ID': row['Source No_'],
                'IMEI': row['Serial No_']
                }

        imei_data.append(imei_dict)

    df3 = pd.DataFrame(imei_data)

    # Create a new Excel workbook-------------------------------------------------------------------------------------------
    wb = openpyxl.Workbook()

    # select the active worksheet
    ws = wb.active
    ws.title = "Summary"

    # write the column headers to the worksheet and format them
    header_font = Font(bold=True)
    header_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    for col_num, col_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_title)
        cell.font = header_font
        cell.border = header_border

    # write the data to the worksheet
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # set the column widths for all columns after data has been written
    for col_num in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # add a new worksheet named "Sales breakdown"-----------------------------------------------------------------------
    ws2 = wb.create_sheet("Sales breakdown")

    # write the column headers to the worksheet and format them
    for col_num, col_title in enumerate(df2.columns, 1):
        cell = ws2.cell(row=1, column=col_num + 1, value=col_title)
        cell.font = header_font
        cell.border = header_border

    # write the data to the worksheet
    for row_num, row_data in enumerate(df2.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws2.cell(row=row_num, column=col_num + 1, value=cell_value)

    # set the column widths for all columns
    for col_num in range(1, ws2.max_column + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws2[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws2.column_dimensions[column_letter].width = adjusted_width

    # add a new worksheet named "Imei Sell thru"------------------------------------------------------------------------
    ws3 = wb.create_sheet("Imei Sell thru")

    # write the column headers to the worksheet and format them
    for col_num, col_title in enumerate(df3.columns, 1):
        cell = ws3.cell(row=2, column=col_num, value=col_title)

    # write the data to the worksheet starting from row 3
    for row_num, row_data in enumerate(df3.values):
        for col_num, cell_value in enumerate(row_data, 1):
            ws3.cell(row=3 + row_num, column=col_num, value=cell_value)

    # set the column widths for all columns after data has been written
    for col_num in range(1, ws3.max_column + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws3[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws3.column_dimensions[column_letter].width = adjusted_width

    # save the workbook-------------------------------------------------------------------------------------------------
    docs_folder = os.path.join(os.environ['USERPROFILE'], 'Documents')
    file_name = f'211000198450_SALES_Inventory_Weekly_{iso_year} W{iso_week}.xlsx'
    full_path = os.path.join(docs_folder, file_name)
    wb.save(full_path)

    # Create the report complete popup----------------------------------------------------------------------------------
    generating_report_popup.destroy()
    report_complete_popup = tk.Toplevel(window)
    report_complete_popup.geometry("250x75")
    report_complete_popup.title("Report Complete")
    report_complete_popup.transient(window)
    report_complete_popup.grab_set()

    report_complete_label = tk.Label(report_complete_popup, text="Report generated successfully!")
    report_complete_label.pack(pady=5)

    def open_file_location():
        os.startfile(os.path.dirname(full_path))

    open_file_location_button = tk.Button(report_complete_popup, text="Open File Location", command=open_file_location)
    open_file_location_button.pack(pady=5)


window = tk.Tk()
window.title("Nokia Sales Report Generator")

# Create the input fields and labels
iso_year_label = tk.Label(window, text="ISO Year:")
iso_year_entry = tk.Entry(window)
iso_week_label = tk.Label(window, text="ISO Week:")
iso_week_entry = tk.Entry(window)
vendor_no_label = tk.Label(window, text="Vendor No:")
vendor_no_entry = tk.Entry(window)
vendor_no_entry.insert(0, "HMDGLOBAL")

# Create generate report button
generate_report_button = tk.Button(window, text="Generate Report", command=lambda: generate_nokia_report(iso_year_entry.get(), iso_week_entry.get(), vendor_no_entry.get()))

# Place the widgets in the window using the grid layout manager
iso_year_label.grid(row=0, column=0, padx=5, pady=5)
iso_year_entry.grid(row=0, column=1, padx=5, pady=5)
iso_week_label.grid(row=1, column=0, padx=5, pady=5)
iso_week_entry.grid(row=1, column=1, padx=5, pady=5)
vendor_no_label.grid(row=2, column=0, padx=5, pady=5)
vendor_no_entry.grid(row=2, column=1, padx=5, pady=5)
generate_report_button.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

# Start the main event loop
window.mainloop()


